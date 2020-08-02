import xlrd
import openpyxl as xlwriter
import os

# move to client directory and load workbooks
os.chdir('/home/prem/temp/client')
book = xlrd.open_workbook('data.xlsx')
same_book = xlwriter.load_workbook('data.xlsx')

#load sheet variables
curr_sheet = book.sheet_by_name(book.sheet_names()[0])
same_sheet = same_book.get_sheet_by_name(book.sheet_names()[0])

#iterate and read value from table rows 1 to n
for i in range(1, curr_sheet.nrows):
    id = int(curr_sheet.cell_value(i, 0))
    name = str(curr_sheet.cell_value(i, 1))
    college_code = str(int(curr_sheet.cell_value(i, 2)))
    course_id = str(int(curr_sheet.cell_value(i, 3)))
    reg_no = str(int(curr_sheet.cell_value(i, 4)))
    year = str(int(curr_sheet.cell_value(i, 5)))
    semester = str(int(curr_sheet.cell_value(i, 6)))
    certificate_version = str(int(curr_sheet.cell_value(i, 7)))
    marks = str(curr_sheet.cell_value(i, 8))
    password = str(curr_sheet.cell_value(i, 9))
    status = curr_sheet.cell_value(i,10)

    # print(id, name, college_code, course_id, reg_no, year, semester, certificate_version, marks, password,status, remarks)
    #execute create command if status of row is active or failed
    if status == "active" or status == "failed":
        command = "node adminCommands createResult {} '{}' {} {} {} {} {} {} '{}' '{}' ".format(id, name, college_code, course_id, reg_no, year, semester, certificate_version, marks, password)
        ret = os.system(command)
        if ret is 256:
            same_sheet.cell(row=i+1, column=11).value = "failed"
            same_book.save('data.xlsx')
        elif ret is 0:
            same_sheet.cell(row=i+1, column=11).value = "inactive"
            same_book.save('data.xlsx')
    else:
        print("skipped row {}".format(i))
    same_book.save('data.xlsx')

#save workbook
same_book.save('data.xlsx')
